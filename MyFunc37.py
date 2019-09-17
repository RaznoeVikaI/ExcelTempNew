import openpyxl

def readWriteCellFilials():
    wb = openpyxl.load_workbook('задача для кандидата.xlsx')
    # получаем имя активного листа
    active_sheet_name = wb.active
    # print(active_sheet_name)
    # получаем другой лист
    sheet_ishod = wb['исходные']
    print(sheet_ishod)

    # кол-во строки в книге
    rows = sheet_ishod.max_row
    # print(rows)

    # номер строки по попрядку
    # index = 1
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         print(str(cell.value))
    #     index +=1

    # вывели все строки первого столбца + если 111 то выводим значение девятого столбца
    # номер строки по попрядку
    # index = 1
    # position_articles = []
    # indexCell = 0
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         # print(str(cell.value))
    #         # print(index)
    #         # position_articles.append(index)
    #         indexCell += 1
    #         if (indexCell == 1):
    #             print(str(cell.value))
    #             # получили значение ячейки "Код статьи" и перевели в int
    #             cellToInt = int(str(cell.value))
    #             if (cellToInt == 111):
    #                 index +=1
    #                 cell = sheet_ishod.cell(row=index, column=9).value
    #                 print(cell)
    #     indexCell = 0
    #     index +=1

    # выводим все значения из ячейки "средний объем"
    # index = 1
    # indexCell = 0
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         # print(str(cell.value))
    #         indexCell += 1
    #         if (indexCell == 9):
    #             print(str(cell.value))
    #             indexCell=0
    #     index +=1

    # индекс равен 1, индекс равен 9
    # index = 1
    # indexCell = 0
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         # print(str(cell.value))
    #         indexCell += 1
    #         if (indexCell == 1):
    #             indexCell+=1
    #             print('indexCell равен 1')
    #         if (indexCell ==9):
    #             print('indexCell равен 9')
    #             # print(str(cell.value))
    #             indexCell=0
    #     index +=1

    # Находит,в стоблце "Код статьи"-111,в стоблце "Сред.обьем" все значения
    # нужно оставить только по выбранному филиалу
    # index = 1
    # indexCell = 0
    # cellToInt = 0
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         indexCell += 1
    #         if (indexCell == 1):
    #             cellToInt = int(cell.value)
    #             if (cellToInt == 111):
    #                 # если нашли такое значение в столбце "Код статьи",то идем и ищем по все строками дальше
    #                 index += 1
    #                 cell = sheet_ishod.cell(row=index, column=9)
    #                 print(cell.value)
    #     indexCell = 0
    #     index += 1

    # # Находим в "Код статьи"-111,потом в "Код филиала"-01,потом по этим параметрам "Средний обьем"
    # # Начинаем обход с первой строки
    # index = 1
    # # Порядковый номер ячейки
    # indexCell = 0
    # while (rows != index):
    #     # смотрим все ячейки строки,номер строки равен index
    #     for cell in list(sheet_ishod.rows)[index]:
    #         indexCell += 1
    #         if (indexCell == 1):
    #             cellToInt = int(cell.value)
    #             # Нужно поле "Код статьи" в UI
    #             if (cellToInt == 111):
    #                 # Для того чтобы убрать строку "Код филиала"
    #                 index +=1
    #                 # Номер филиала: 01, 02,...
    #                 cell = sheet_ishod.cell(row=index, column=5)
    #                 # убираем ноль, чтобы можно было сравнивать
    #                 print(cell.value[1:])
    #                 cellToInt = int(cell.value[1:])
    #                 # Нужно поле "Код филиала" в UI
    #                 if (cellToInt == 1):
    #                     cell = sheet_ishod.cell(row=index, column=9)
    #                     print(cell.value)
    #                     # если нашли такое значение в столбце "Код статьи",то идем и ищем по все строками дальше
    #                     index += 1
    #     indexCell = 0
    #     index += 1

    # Находим в "Код статьи"-111,потом в "Код филиала"-01,потом по этим параметрам берем "Средний обьем",потом сумму
    # Начинаем обход с первой строки
    index = 1
    # Порядковый номер ячейки
    indexCell = 0
    # Сумма среднего объема по коду статьи и филиалу
    sredObyiemSum_01 = 0
    sredObyiemSum_02 = 0
    sredObyiemSum_03 = 0
    # колличество найденных элментов - тоесть 1)по статье 2)по филиалу
    countArticleFilial_01 = 0
    countArticleFilial_02 = 0
    countArticleFilial_03 = 0
    # Результаты вычислений - Среднеквартальный обьем
    volumeIshod = {}
    # колличество найденных элментов - тоесть 1)по статье 2)по филиалу
    countArticleFilial = 0
    # Результаты вычислений - Среднеквартальный обьем
    volumeIshod = {}
    while (rows != index):
        # смотрим все ячейки строки,номер строки равен index
        for cell in list(sheet_ishod.rows)[index]:
            indexCell += 1
            if (indexCell == 1):
                cellToInt = int(cell.value)
                # Нужно поле "Код статьи" в UI
                if (cellToInt == 111):
                    # Для того чтобы убрать строку "Код филиала"
                    index += 1
                    # Номер филиала: 01, 02,...
                    cell = sheet_ishod.cell(row=index, column=5)
                    # убираем ноль, чтобы можно было сравнивать
                    # print(cell.value[1:])
                    cellToInt = int(cell.value[1:])
                    # Нужно поле "Код филиала" в UI
                    if (cellToInt == 1):
                        codeFilial = cellToInt
                        countArticleFilial +=1
                        cell = sheet_ishod.cell(row=index, column=9)
                        # print(countArticleFilial)
                        # print(cell.value)
                        sredObyiemSum_01 +=int(cell.value)
                        volumeIshod.update({codeFilial: sredObyiemSum_01})
                        # если нашли такое значение в столбце "Код статьи",то идем и ищем по все строками дальше
                        index += 1
        indexCell = 0
        index += 1
    sredObyiemSum_01 = sredObyiemSum_01 / countArticleFilial
    # articlesStavki.update({stavkiKey: indexStavki})
    # volumeIshod.update({cellToInt:sredObyiemSum})
    print(sredObyiemSum_01)

    # Пока посчитали только по 111 и код филиала-01

    # Пытаемся писать
    print("МЕНЯЕМ АКТИВНЫЙ ЛИСТ ДЛЯ ЗАПИСИ")
    sheet_otchet = wb.get_sheet_by_name('отчет')

    # for cell in sheet_otchet['A']:
    #     print(cell.value)

    # ----------111---------------
    # создаем список ключей,чтобы потом перебирать
    keysList = list(volumeIshod.keys())
    # print("список ключей")
    # for i in keysList:
    #     print(i)
    #     print(type(i))

    # выводим значение ячейки "А" из листа "отчет"
    # indexRowOtchet = 0
    # for cell in sheet_otchet['A']:
    #     indexRowOtchet += 1
    #     if (indexRowOtchet >= 6):
    #         cellToInt = int(cell.value)
    #         if (cellToInt == 111):
    #             print("Значения ячейки А")
    #             # одно значение выводится,так как пока ищем только 111
    #             print(cell.value)

    print("ПЫТАЕМСЯ ПИСАТЬ")
    # print("sredObyiemSum")
    # print(sredObyiemSum_01)
    indexRowOtchet = 0
    for cell in sheet_otchet['A']:
        indexRowOtchet += 1
        if (indexRowOtchet >= 6):
            cellToInt = int(cell.value)
            # int
            # print(type(cellToInt))
            if (cellToInt == 111):
                if (indexRowOtchet == 7):
                    for i in keysList:
                        # print("Ключи с расчитаныыми объемами")
                        # print(i)
                        # в зависимости от выбора "номер филиала" заполняются ячейки напротив "код статьи"
                        # расчитанными средними объемами
                        if (i == 1):
                            print("Значения ячейки А i == 1")
                            # одно значение выводится,так как пока ищем только 111
                            # print(cellToInt)
                            # E +indexRowOtchet = E7
                            cellAddress = "E" + str(indexRowOtchet)
                            print("sredObyiemSum_01")
                            print(sredObyiemSum_01)
                            sheet_otchet[cellAddress].value = sredObyiemSum_01
                        # в зависимости от выбора "номер филиала" заполняются ячейки напротив "код статьи"
                        # расчитанными средними объемами
                        if (i == 2):
                            print("Значения ячейки А i == 2")
                            # одно значение выводится,так как пока ищем только 111
                            # print(cell.value)
                            # print("Ключи с расчитаныыми объемами")
                            # print(i)
                            # клеим
                            # E +indexRowOtchet = E7
                            # cellAddress = "G" + str(indexRowOtchet)
                            # sheet_otchet[cellAddress].value = volumeIshod[i] / countArticleFilial_02
                            # в зависимости от выбора "номер филиала" заполняются ячейки напротив "код статьи"
                            # расчитанными средними объемами
                        if (i == 3):
                            print("Значения ячейки А i == 3")
                            # одно значение выводится,так как пока ищем только 111
                            # print(cell.value)
                            # print("Ключи с расчитаныыми объемами")
                            # print(i)
                            # клеим
                            # E +indexRowOtchet = E7
                            # cellAddress = "G" + str(indexRowOtchet)
                            # sheet_otchet[cellAddress].value = volumeIshod[i] / countArticleFilial_03
    indexRowOtchet += 1
    # Попробуем сохранить файл
    wb.save('задача для кандидата.xlsx')

    # print("volumeIshod")
    # 1 -> 11238
    # 2 -> 13725
    # 3 -> 14941
    # for key in volumeIshod:
    #     print("%s -> %s" % (key, volumeIshod[key]))

    # получаем ячейку E5 листа "отчет"
    # просто проверяем связь
    # cell = sheet_otchet.cell(row=4, column=5)
    # print(cell.value)

    # --------111------------