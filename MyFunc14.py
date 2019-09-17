import openpyxl

def readWriteCellFilials():
    wb = openpyxl.load_workbook('задача для кандидата.xlsx')
    # получаем имя активного листа
    active_sheet_name = wb.active
    # print(active_sheet_name)
    # получаем другой лист
    sheet_ishod = wb['исходные']
    # print(sheet_ishod)

    # печатаем значение ячейки F1
    # print(sheet_ishod['F1'].value)

    # получаем ячейку листа A2
    # cell = sheet_ishod['A2']
    # 2
    # print('Строка: ' + str(cell.row))
    # A2
    # print('Ячейка: ' + cell.coordinate)
    # 111
    # print('Значение: ' + cell.value)

    # получаем ячейку A2 листа исходные
    # cell = sheet_ishod.cell(row=2, column=1)
    # print(cell.value)

    # выводим содержимое всего листа
    # rows = sheet_ishod.max_row
    # cols = sheet_ishod.max_column
    #
    # for i in range(1, rows + 1):
    #     string = ''
    #     for j in range(1, cols + 1):
    #         cell = sheet_ishod.cell(row=i, column=j)
    #         string = string + str(cell.value) + ' '
    #     print(string)

    # выводим содержимое определенного диапазона листа
    # for row in sheet_ishod['A1':'C3']:
    #     string = ''
    #     for cell in row:
    #         string = string + str(cell.value) + ' '
    #     print(string)

    # выводим содержимое столбика
    # for cell in sheet_ishod['A']:
    #     print(cell.value)

    # получаем другой лист
    sheet_spiski = wb['списки']
    # print(sheet_spiski)

    # Список для хранения значений из столбца "Название филиала"
    names_filials = []
    # Print out values in column "Название филиала"
    for i in range(2, 5):
        names_filials.append(sheet_spiski.cell(row=i, column=2).value)

    # names_filials - выводит всю строку
    print("MyFunc14")
    # print(names_filials)

    # names_filials - выводит поэлементно
    # for i in range(len(names_filials)):
    #     print(names_filials[i])

    # выводим один элмент
    # print(names_filials[0])

    # получаем другой лист
    sheet_ishod = wb['исходные']
    # print(sheet_ishod)

    # Пишем в столбик "Название филиала" листа "Исходные" из столбика "Название филиала" листа "Списки"
    # По кол-ву элементов в списке
    countList = len(names_filials);
    # для того чтобы правильно отсчитывать строки
    rowSheet = 2
    # для прохождения с начала списка names_filials
    rowList = 0

    while countList > 0:
        # выводим содержимое списка names_filials
        # print(names_filials[rowList])

        # выводим содержимео одной ячейки
        # print(sheet_ishod.cell(row=2,column=3).value)

        # выводим содержание столбика "Название филиала" книги "Списки"
        # print(sheet_spiski.cell(row=rowSheet,column=2).value)

        # присваеиваем строкам стоблца "Название филиала" в книге "Исходные"
        sheet_ishod.cell(row=rowSheet, column=6).value = names_filials[rowList]
        rowList +=1
        countList-=1
        rowSheet +=1

    # Попробуем сохранить файл
    wb.save('задача для кандидата.xlsx')

    # Выводим содержимое всего листа после записи данных в колонку "Название филиала"
    rows = sheet_ishod.max_row
    cols = sheet_ishod.max_column

    for i in range(1, rows + 1):
        string = ''
        for j in range(1, cols + 1):
            cell = sheet_ishod.cell(row=i, column=j)
            string = string + str(cell.value) + ' '
        print(string)