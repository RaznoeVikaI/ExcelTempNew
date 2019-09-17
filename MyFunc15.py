import openpyxl

def readWriteCellFilials():
    wb = openpyxl.load_workbook('задача для кандидата.xlsx')
    # получаем имя активного листа
    active_sheet_name = wb.active
    # print(active_sheet_name)
    # получаем другой лист
    sheet_ishod = wb['исходные']
    # print(sheet_ishod)

    # получаем другой лист
    sheet_spiski = wb['списки']
    # print(sheet_spiski)

    # Список для хранения значений из столбца "Название статьи 1ур"
    names_articles_1 = []
    # Print out values in column "Название филиала"
    for i in range(2, 10):
        names_articles_1.append(sheet_spiski.cell(row=i, column=6).value)

    # print(names_articles_1)
    print("MyFunc15")

    # получаем другой лист
    sheet_ishod = wb['исходные']
    # print(sheet_ishod)

    # Пишем в столбик "Название статьи 1ур" листа "Исходные" из столбика "Название статьи 1ур" листа "Списки"
    # По кол-ву элементов в списке
    countList = len(names_articles_1);
    # для того чтобы правильно отсчитывать строки
    rowSheet = 2
    # для прохождения с начала списка names_filials
    rowList = 0

    while countList > 0:
        # выводим содержимое списка names_filials
        # print(names_articles_1[rowList])

        # выводим содержимео одной ячейки
        # print(sheet_ishod.cell(row=2,column=3).value)

        # выводим содержание столбика "Название филиала" книги "Списки"
        # print(sheet_spiski.cell(row=rowSheet,column=2).value)

        # присваеиваем строкам стоблца "Название филиала" в книге "Исходные"
        sheet_ishod.cell(row=rowSheet, column=2).value = names_articles_1[rowList]
        rowList += 1
        countList -= 1
        rowSheet += 1

    # Выводим содержимое всего листа после записи данных в колонку "Название филиала"
    rows = sheet_ishod.max_row
    cols = sheet_ishod.max_column

    for i in range(1, rows + 1):
        string = ''
        for j in range(1, cols + 1):
            cell = sheet_ishod.cell(row=i, column=j)
            string = string + str(cell.value) + ' '
        print(string)