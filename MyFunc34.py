import openpyxl

def readWriteCellFilials():
    stavkiKey = 0

    wb = openpyxl.load_workbook('задача для кандидата.xlsx')
    # получаем имя активного листа
    active_sheet_name = wb.active
    # print(active_sheet_name)
    # получаем другой лист
    # пункт 1 - Открываем нужный лист
    sheet_stavki = wb['ставки']
    print(sheet_stavki)

    # кол-во строки в книге
    rows = sheet_stavki.max_row
    # print(rows)
    # Начинаем обход с третьей строки
    index = 2
    # среди трех колонок  нам нужна вторая и третья там формулы
    indexColumns = 1
    # Процентные ставки по статье.Сначала копим,потом кладем их в словарь
    stavkiArticle = []
    stavkiValue = 0
    # Беру проц.ставку за месяц делю на 12, три раза, по каждой статье - Средняя ставка за три месяца
    stavkiOctNovDec = 0
    # Среднюю ставку за три месяца умножим на 4(четыре квартала за год)-получим годовую процентную ставку
    stavkaYear = 0
    # Статьи - по этим ставкам будем искать в другом листе статьи и записывать результа вычислений - годовой процентной ставки
    articlesStavki = {}
    # Порядковый номер ставки
    indexStavki = 1
    # Результаты вычислений годовых процентов
    procentStavki = {}
    # Порядковый номер вычесленных годовых процентов
    indexProcent = 0
    while (rows != index):
        # смотрим все ячейки строки,номер строки равен index
        # пункт 2 - Проходим по всем строкам и столбцам листа
        for cell in list(sheet_stavki.rows)[index]:
            # пункт 2.1 - первый столбик - нам нужны проценты по каждой статье
            if (indexColumns == 1):
                # получили первый ключ
                stavkiKey = cell.value
                # 111, например
                # print(stavkiKey)
                articlesStavki.update({stavkiKey: indexStavki})
                # articlesStavki.update({indexStavki:stavkiKey})
                indexStavki += 1
                indexProcent +=1
            # первый столбик со ставками
            if (indexColumns == 2):
                # пункт 2.1 - получаем сами проценты.Только внимание:
                # первый процент возвращается цифрой 0.2
                stavkiValue = cell.value
                # 0,2 например
                # print("второй столбик")
                # print(stavkiValue)
                stavkiArticle.append(stavkiValue)
            # второй столбик со ставками
            # ВЫВОД: ПО КАЖДДОЙ СТАТЬЕ НУЖНО ЗНАТЬ СТАВКУ И НАДБАВКУ - МОЖЕТ ЭТО ПОЛЯ ДЛЯ ui
            if (indexColumns == 3):
                # пункт 2.2 - получаем сами проценты.Только внимание:
                # второй,третий процент возвращается формулой: =B3*1.02
                # 0.2*1.02=0.204
                stavkiValue = stavkiValue*1.02
                # округлили до трех знаков после запятой.
                # В Python есть особенность-здесь используется "Банковское округление",
                # то есть округление к ближайшему чётному
                stavkiArticle.append(stavkiValue)
            if (indexColumns == 4):
                # пункт 2.3 - получаем сами проценты.Только внимание:
                # второй,третий процент возвращается формулой: =C3*1.02
                # 0.204*1.02=0.20808
                stavkiValue = stavkiValue*1.02
                # Округляем stavkiValue используя "Банковское округление"
                stavkiValue = round(stavkiValue,3)
                # пункт 3 - Кладем полученные процентные в список
                stavkiArticle.append(stavkiValue)
            indexColumns +=1
            # print(stavkiKey)
        index += 1
        indexColumns=1

        # Пытаемся писать
        print("МЕНЯЕМ АКТИВНЫЙ ЛИСТ ДЛЯ ЗАПИСИ")
        sheet_otchet = wb.get_sheet_by_name('отчет')
        print(sheet_otchet)
        indexColumnOtchet =1
        for cell in sheet_otchet['A']:
            if (indexColumnOtchet >=7):
                # print(cell.value)
                cellToInt = int(cell.value)
                # НУЖНО ПОДУМАТЬ КАК ПЕРЕБРАТЬ ВСЕ СТАТЬИ, может забить в файл или поле с введенными статьями на UI
                if (cellToInt == 111
                        or cellToInt == 112
                        or cellToInt == 121
                        or cellToInt == 122
                        or cellToInt == 211
                        or cellToInt == 212
                        or cellToInt == 221
                        or cellToInt == 222):
                    print(cellToInt)
                    print("Номер строки:" + str(indexColumnOtchet))
                    # Процентная ставка
                    # j=8
                    # sheet_otchet.cell(row=indexColumnOtchet-1, column=j).value = str(cellToInt)
                    print("RRR")
                    for key in articlesStavki:
                        # print("%s -> %s" % (key, articlesStavki[key]))
                        # по ключу-111 ищем значение - это 1
                        if (key == str(cellToInt)):
                            print("find")
                            # print("%s -> %s" % (key, articlesStavki[key]))
                            print(articlesStavki[key])
                            keyArticlesStavki = articlesStavki[key]
                            # после того как по значение мы вытащили(нашли) ключ, идем во второй словарь
                            # по вытащенному ключу вытаскиваем значение во втором словаре
                            for key in procentStavki:
                                if(key == keyArticlesStavki):
                                    print('find2')
                                    # print(procentStavki[key])
                                    procentStavkiOkrugl = round(procentStavki[key], 2)
                                    # print(procentStavkiOkrugl)
                                    sheet_otchet.cell(row=indexColumnOtchet,column=6).value = procentStavkiOkrugl
                                    # print(indexColumnOtchet)
            indexColumnOtchet +=1
        # Пытаемся писать
        print(stavkiKey)
        # пункт4 - Берем из списка,возвращаем в проценты обратно и делим на 12 -
        # чтобы узнать вклад этого процента в годовой процент,после перевода каждого складываем
        for x in stavkiArticle:
            # print(x)  # выводит вертикально
            # пункт 3 -Делаем для трех месяцев каждой статьи
            # пункт 4 - После складываем
            stavkiOctNovDec += (x * 100)/12
        # пункт 5 - ГОДОВАЯ СТАВКА ПО КАЖДОЙ СТАТЬЕ = У нас четере квартала, поэтому умножаем на 4
        stavkaYear = stavkiOctNovDec * 4
        # ГОДОВАЯ СТАВКА ПО СТАТЬЕ
        print(stavkaYear)
        # Пытаемся писать
        # print(sheet_otchet['F5'].value)

        # Процентная ставка
        # j = 6
        # sheet_otchet.cell(row=indexColumnOtchet - 1, column=j).value = str(stavkaYear)
        # sheet_otchet.cell(row=5, column=6).value = str(stavkaYear)
        # Второй словарь: порядоквый номер(ключ) - годовая ставка(значение)
        # procentStavki.update({stavkaYear: indexProcent})
        procentStavki.update({indexProcent:stavkaYear})

        # Поиск в словаре нужного значения,затем получаем ключ и идем во второй словарь
        # for key in articlesStavki:
        #     if (articlesStavki[key] == "111"):
        #         print("find")
        # Поиск в словаре нужного значения,затем получаем ключ и идем во второй словарь

        # Попробуем сохранить файл
        # wb.save('задача для кандидата.xlsx')
        # Пытаемся писать
        # для следующей статьи нужно все переменные очистить и проделать все с пункта 1
        stavkiArticle.clear()
        stavkiOctNovDec = 0
        stavkaYear = 0

        # # Попробуем сохранить файл
        wb.save('задача для кандидата.xlsx')

        print("DICT-1")
        for key in articlesStavki:
            print("%s -> %s" % (key, articlesStavki[key]))

        print("DICT-2")
        for key in procentStavki:
            print("%s -> %s" % (key, procentStavki[key]))