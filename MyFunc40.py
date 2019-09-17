import openpyxl

def readWriteCellFilials():
    wb = openpyxl.load_workbook('задача для кандидата.xlsx')
    # получаем имя активного листа
    active_sheet_name = wb.active
    # print(active_sheet_name)
    # получаем другой лист
    sheet_otchet = wb['отчет']
    print(sheet_otchet)

    # ----------111------------
    # итоговая процентная маржа по всем статьям
    SummaryMargin = 0
    # заглушка
    i =1
    print("ПЫТАЕМСЯ ПИСАТЬ")
    indexRowOtchet = 0
    for cell in sheet_otchet['A']:
        indexRowOtchet += 1
        if (indexRowOtchet >= 7):
            cellToInt = int(cell.value)
            # int
            # print(type(cellToInt))
            if (cellToInt == 111):
                if (indexRowOtchet == 7):
                    # for i in keysList:
                        # print("Ключи с расчитаныыми объемами")
                        # print(i)
                        # в зависимости от выбора "номер филиала" расчитывается ячейка "процентные доходы-расходы"
                        # может быть тоже не нужно, т.к. мы ведь берем ячейки уже посчитанные и делаем с ними
                        # вычисления дальше
                        if (i == 1):
                            print("Значения ячейки А i == 1")
                            cellA_Address = "A" + str(indexRowOtchet)
                            print(sheet_otchet[cellA_Address].value)
                            # Первый операнд для вычисления процентной маржи
                            cellE_Address = "E" + str(indexRowOtchet)
                            # 3746
                            print(sheet_otchet[cellE_Address].value)
                            cellG_Address = "G" + str(indexRowOtchet)
                            # 764.184
                            print(sheet_otchet[cellG_Address].value)
                            #  Результат вычисления маржи пишем в строку следующую за последней со статьями
                            print(sheet_otchet.max_row)
                            lastArticlesRows = sheet_otchet.max_row
                            cellG_Margin_Address = "G" + str(lastArticlesRows+1)
                            ResultMargin = ((sheet_otchet[cellE_Address].value - sheet_otchet[cellG_Address].value)/\
                                           sheet_otchet[cellE_Address].value)*100
                            SummaryMargin +=ResultMargin
                            sheet_otchet[cellG_Margin_Address].value = SummaryMargin


    indexRowOtchet += 1
    # Попробуем сохранить файл
    wb.save('задача для кандидата.xlsx')
    # ----------111------------